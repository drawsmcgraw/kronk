"""Format-agnostic positions importer ("Kronk must be clever about parsing").

Two-stage design (docs/plans/FINANCIAL_EXPERT_PLAN.md):
1. Mapping discovery — once per source format. An LLM sees headers + a few
   sample rows and proposes which column is holding/value/shares/price.
   The mapping (not the data) is the model's only output, and the operator
   confirms it before first use.
2. Deterministic extraction — every import. Code applies the stored mapping
   to every row and enforces invariants BEFORE anything lands:
     * every mapped cell parses (money/number/date),
     * shares x price ~= value (1%) wherever all three exist,
     * a stated total row must reconcile with the sum of rows (0.5%),
     * an as-of date must be found (mapping column, filename, or supplied).
   Any violation rejects the whole import with row-level detail — no
   partial ingests, no OCR guessing, no model-transcribed numbers.

Supported formats: CSV/TSV, XLSX (openpyxl), JSON (list of objects).
PDF/OFX are rejected with "export CSV/XLSX instead" — deliberately.
"""
import csv
import io
import json
import logging
import os
import re
from dataclasses import dataclass, field

import httpx

logger = logging.getLogger(__name__)

LLM_URL       = os.getenv("LLM_SERVICE_URL", "http://host.docker.internal:8002")
MAPPING_MODEL = os.getenv("MAPPING_MODEL", "gemma-4-e4b")

REQUIRED_FIELDS = ("holding", "value")
OPTIONAL_FIELDS = ("shares", "price", "as_of_date", "account")
# account_map: when the file has an account column (brokerage exports often
# cover several accounts in one file), mapping["account_map"] translates
# each distinct column value to a kronk account id — or "__skip__".
SKIP_ACCOUNT = "__skip__"
_TOTAL_ROW_RE   = re.compile(r"^\s*(grand\s+)?total\b", re.IGNORECASE)
_FILENAME_DATE_RE = re.compile(r"(20\d{2})[-_.]?(\d{2})[-_.]?(\d{2})")
# US brokerage statement convention: Statement06302026.csv (MMDDYYYY).
_FILENAME_DATE_US_RE = re.compile(r"(?<!\d)(\d{2})(\d{2})(20\d{2})(?!\d)")


class ImportError422(Exception):
    """Import rejected — carries the specific, user-facing reason."""


@dataclass
class ParsedFile:
    format: str
    headers: list[str]
    rows: list[dict]                # header -> raw cell (str)
    sample: list[dict] = field(default_factory=list)


# ── parsing primitives ────────────────────────────────────────────────────────

def parse_money(raw) -> float:
    """'$1,234.56' → 1234.56; '(123.45)' → -123.45. Raises ValueError."""
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace("$", "").replace(",", "").replace("%", "")
    if not s or s in ("-", "--", "N/A"):
        raise ValueError(f"empty/non-numeric cell: {raw!r}")
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    val = float(s)
    return -val if neg else val


def parse_date_cell(raw) -> str:
    """Normalize common date shapes to YYYY-MM-DD. Raises ValueError."""
    s = str(raw).strip()[:10]
    for pattern, order in ((r"^(\d{4})-(\d{2})-(\d{2})", (1, 2, 3)),
                           (r"^(\d{1,2})/(\d{1,2})/(\d{4})", (3, 1, 2))):
        m = re.match(pattern, s)
        if m:
            y, mo, d = (m.group(order[0]), m.group(order[1]), m.group(order[2]))
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    raise ValueError(f"unrecognized date: {raw!r}")


def sniff_and_parse(data: bytes, filename: str) -> ParsedFile:
    name = filename.lower()
    if name.endswith((".pdf", ".ofx", ".qfx")):
        raise ImportError422(
            f"{name.rsplit('.', 1)[-1].upper()} imports are not supported — "
            "export CSV or XLSX from the brokerage instead (we don't guess "
            "at money).")
    if name.endswith(".xlsx"):
        return _parse_xlsx(data)
    if name.endswith(".json"):
        return _parse_json(data)
    # CSV/TSV (default): sniff the delimiter.
    return _parse_csv(data)


def _parse_csv(data: bytes) -> ParsedFile:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    # Brokerage CSVs often carry preamble lines ("Account: ...") before the
    # header. Find the first line that yields >= 2 columns and treat it as
    # the header row.
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        raise ImportError422("file is empty")
    delimiter = "\t" if lines[0].count("\t") > lines[0].count(",") else ","
    start = 0
    for i, ln in enumerate(lines):
        if len([c for c in next(csv.reader([ln], delimiter=delimiter)) if c.strip()]) >= 2:
            start = i
            break
    reader = csv.DictReader(io.StringIO("\n".join(lines[start:])), delimiter=delimiter)
    headers = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
    if len(headers) < 2:
        raise ImportError422("could not find a header row with at least two columns")
    rows = [{k.strip(): (v or "").strip() for k, v in row.items() if k}
            for row in reader]
    rows = [r for r in rows if any(v for v in r.values())]
    return ParsedFile("csv", headers, rows, rows[:3])


def _parse_xlsx(data: bytes) -> ParsedFile:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError422("XLSX support not installed (openpyxl missing)")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ImportError422(f"could not open XLSX: {e}")
    ws = wb.worksheets[0]
    all_rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    header_idx = next((i for i, r in enumerate(all_rows)
                       if sum(1 for c in r if isinstance(c, str) and c.strip()) >= 2),
                      None)
    if header_idx is None:
        raise ImportError422("could not find a header row in the first sheet")
    headers = [str(c).strip() for c in all_rows[header_idx] if c is not None and str(c).strip()]
    rows = []
    for r in all_rows[header_idx + 1:]:
        cells = {headers[i]: ("" if r[i] is None else r[i])
                 for i in range(min(len(headers), len(r)))}
        if any(str(v).strip() for v in cells.values()):
            rows.append(cells)
    return ParsedFile("xlsx", headers, rows, rows[:3])


def _parse_json(data: bytes) -> ParsedFile:
    try:
        obj = json.loads(data)
    except ValueError as e:
        raise ImportError422(f"invalid JSON: {e}")
    if not isinstance(obj, list) or not obj or not isinstance(obj[0], dict):
        raise ImportError422("JSON must be a non-empty array of objects")
    headers = sorted({k for row in obj for k in row})
    return ParsedFile("json", headers, obj, obj[:3])


def fingerprint(headers: list[str]) -> str:
    """Stable id for a source format: order- and case-insensitive."""
    import hashlib
    canon = "|".join(sorted(h.strip().lower() for h in headers))
    return hashlib.sha256(canon.encode()).hexdigest()[:16]


# ── mapping application + invariants ──────────────────────────────────────────

def validate_mapping(mapping: dict, headers: list[str]) -> list[str]:
    """Structural check — used on LLM proposals AND operator submissions."""
    errors = []
    cols = mapping.get("columns") or {}
    for f in REQUIRED_FIELDS:
        if not cols.get(f):
            errors.append(f"mapping is missing required field '{f}'")
    lower_headers = {h.lower() for h in headers}
    for f, col in cols.items():
        if col and str(col).lower() not in lower_headers:
            errors.append(f"mapped column {col!r} (for '{f}') is not in the file headers")
        if f not in REQUIRED_FIELDS + OPTIONAL_FIELDS:
            errors.append(f"unknown mapped field {f!r}")
    return errors


def apply_mapping(parsed: ParsedFile, mapping: dict,
                  as_of_override: str | None, filename: str
                  ) -> tuple[dict[str | None, list[dict]], str]:
    """Deterministic extraction. Returns (groups, as_of_date) where groups
    maps the account-column value (None when the file has no account
    column) to canonical rows. Raises ImportError422 with row-level detail
    on any invariant failure."""
    cols = mapping["columns"]
    account_col = cols.get("account")
    # Header lookup is case-insensitive (fingerprint already is).
    def col_key(row: dict, name: str):
        for k in row:
            if k.lower() == str(name).lower():
                return row[k]
        return ""

    groups: dict[str | None, list[dict]] = {}
    errors: list[str] = []
    stated_total: float | None = None
    dates: set[str] = set()

    for i, row in enumerate(parsed.rows, start=1):
        holding = str(col_key(row, cols["holding"])).strip()
        raw_value = col_key(row, cols["value"])
        if not holding and not str(raw_value).strip():
            continue  # fully blank line
        if _TOTAL_ROW_RE.match(holding):
            try:
                stated_total = parse_money(raw_value)
            except ValueError:
                pass
            continue
        entry: dict = {"holding": holding}
        try:
            entry["value"] = parse_money(raw_value)
            for f in ("shares", "price"):
                if cols.get(f) and str(col_key(row, cols[f])).strip():
                    entry[f] = parse_money(col_key(row, cols[f]))
            if cols.get("as_of_date") and str(col_key(row, cols["as_of_date"])).strip():
                dates.add(parse_date_cell(col_key(row, cols["as_of_date"])))
        except ValueError as e:
            errors.append(f"row {i} ({holding[:40] or 'no name'}): {e}")
            continue
        if not holding:
            errors.append(f"row {i}: has a value but no holding name")
            continue
        # shares x price ~= value where all three exist (1% tolerance)
        if "shares" in entry and "price" in entry and entry["value"]:
            calc = entry["shares"] * entry["price"]
            if abs(calc - entry["value"]) > max(abs(entry["value"]) * 0.01, 0.02):
                errors.append(
                    f"row {i} ({holding[:40]}): shares x price = {calc:,.2f} "
                    f"but value = {entry['value']:,.2f}")
                continue
        acct_val = (str(col_key(row, account_col)).strip() or None) if account_col else None
        groups.setdefault(acct_val, []).append(entry)

    if errors:
        shown = "; ".join(errors[:5])
        more = f" (+{len(errors) - 5} more)" if len(errors) > 5 else ""
        raise ImportError422(f"import rejected — {len(errors)} bad row(s): {shown}{more}")
    if not groups:
        raise ImportError422("import rejected — no data rows found after mapping")

    total = sum(r["value"] for rows in groups.values() for r in rows)
    if stated_total is not None and abs(total - stated_total) > max(abs(stated_total) * 0.005, 0.02):
        raise ImportError422(
            f"import rejected — file's stated total {stated_total:,.2f} does not "
            f"match the sum of rows {total:,.2f}")

    # as-of date: explicit override > in-file column > filename.
    as_of = as_of_override
    if not as_of and dates:
        if len(dates) > 1:
            raise ImportError422(
                f"import rejected — multiple as-of dates in one file: {sorted(dates)}")
        as_of = dates.pop()
    if not as_of:
        m = _FILENAME_DATE_RE.search(filename)
        if m:
            as_of = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        else:
            m = _FILENAME_DATE_US_RE.search(filename)
            if m and 1 <= int(m.group(1)) <= 12 and 1 <= int(m.group(2)) <= 31:
                as_of = f"{m.group(3)}-{m.group(1)}-{m.group(2)}"
    if not as_of:
        raise ImportError422(
            "no as-of date found — pass ?as_of=YYYY-MM-DD, or include a date "
            "column / a date in the filename")
    return groups, as_of


# ── LLM mapping proposal (once per new format) ────────────────────────────────

async def propose_mapping(headers: list[str], sample: list[dict]) -> dict | None:
    """Ask the LLM which columns map to the canonical fields. Returns a
    structurally-validated mapping dict, or None (caller falls back to
    manual mapping — ingestion never hard-depends on a model)."""
    prompt = (
        "You map spreadsheet columns from a brokerage positions export to a "
        "canonical schema. Respond with ONLY a JSON object, no prose:\n"
        '{"columns": {"holding": "<column>", "value": "<column>", '
        '"shares": "<column or null>", "price": "<column or null>", '
        '"as_of_date": "<column or null>", "account": "<column or null>"}}\n'
        "- holding: the fund/ticker/position NAME column\n"
        "- value: the current market value in dollars (NOT cost basis, NOT "
        "gain/loss)\n"
        "- account: the column identifying WHICH account each row belongs "
        "to, if the file covers several accounts (account name/number)\n\n"
        f"Headers: {json.dumps(headers)}\n"
        f"Sample rows: {json.dumps(sample[:3], default=str)[:1500]}"
    )
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                f"{LLM_URL}/v1/chat/completions",
                json={"model": MAPPING_MODEL, "temperature": 0,
                      "messages": [{"role": "user", "content": prompt}]},
            )
        if resp.status_code != 200:
            logger.warning("mapping proposal: LLM returned %s: %s",
                           resp.status_code, resp.text[:200])
            return None
        content = resp.json()["choices"][0]["message"]["content"] or ""
        m = re.search(r"\{.*\}", content, re.DOTALL)
        if not m:
            logger.warning("mapping proposal: no JSON in response: %r", content[:200])
            return None
        mapping = {"columns": {k: v for k, v in
                               (json.loads(m.group(0)).get("columns") or {}).items()
                               if v}}
    except Exception as e:
        logger.warning("mapping proposal failed: %s", e)
        return None
    if validate_mapping(mapping, headers):
        logger.warning("mapping proposal failed validation: %s "
                       "(headers=%s)", mapping, headers)
        return None
    return mapping
